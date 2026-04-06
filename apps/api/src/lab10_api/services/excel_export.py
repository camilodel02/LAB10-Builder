"""Genera Excel consolidado alineado a columnas A–P del brief."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Encabezados según plantilla-brief-ia.md
EXCEL_HEADERS: list[str] = [
    "Comprobante No.",
    "NIT_Proveedor",
    "Razon_Social",
    "Fecha_Emision",
    "Numero_Factura",
    "Concepto_Gasto",
    "Monto_COP",
    "Monto_COP_Letras",
    "Confianza_Extraccion_IA (%)",
    "Estado_Validacion_IA",
    "Aprobado_Por",
    "Fecha_Aprobacion",
    "Centro_Costo",
    "Observaciones",
    "Fecha_Carga_Sistema",
    "Ciudad",
]


def _cell_value(v: Any) -> Any:
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def build_expense_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=_cell_value(row.get("id")))
        ws.cell(row=r_idx, column=2, value=row.get("nit_proveedor"))
        ws.cell(row=r_idx, column=3, value=row.get("razon_social"))
        ws.cell(row=r_idx, column=4, value=_cell_value(row.get("fecha_emision")))
        ws.cell(row=r_idx, column=5, value=row.get("numero_factura"))
        ws.cell(row=r_idx, column=6, value=row.get("concepto_gasto"))
        ws.cell(row=r_idx, column=7, value=_cell_value(row.get("monto_cop")))
        ws.cell(row=r_idx, column=8, value=row.get("monto_cop_letras") or "")
        ws.cell(row=r_idx, column=9, value=_cell_value(row.get("confianza_extraccion_ia")))
        ws.cell(row=r_idx, column=10, value=row.get("estado_validacion_ia"))
        ws.cell(row=r_idx, column=11, value=_cell_value(row.get("aprobado_por")))
        ws.cell(row=r_idx, column=12, value=_cell_value(row.get("fecha_aprobacion")))
        ws.cell(row=r_idx, column=13, value=row.get("centro_costo"))
        ws.cell(row=r_idx, column=14, value=row.get("observaciones"))
        ws.cell(row=r_idx, column=15, value=_cell_value(row.get("fecha_carga_sistema")))
        ws.cell(row=r_idx, column=16, value=row.get("ciudad"))
    for col in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
